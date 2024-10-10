import unittest
from unittest.mock import patch, MagicMock
import openpyxl
import sys 
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from src.prep.renaming import rename_sheets

# from file_utils import rename_sheets

class TestRenameSheets(unittest.TestCase):
    @patch('openpyxl.load_workbook')
    def test_rename_sheets(self, mock_load_workbook):
        # Create a mock workbook object
        mock_workbook = MagicMock()
        mock_workbook.sheetnames = ['Sheet1', 'Sheet2']

        # Create mock sheets within the workbook
        mock_sheet1 = MagicMock()
        mock_sheet2 = MagicMock()

        # Assign the mock sheets to the workbook's sheetnames
        mock_workbook.__getitem__.side_effect = [mock_sheet1, mock_sheet2]

        # Mock load_workbook to return the mock workbook
        mock_load_workbook.return_value = mock_workbook

        # Define the sheet mapping
        sheet_mapping = {'Sheet1': 'RenamedSheet1', 'Sheet2': 'RenamedSheet2'}

        # Call the function
        rename_sheets('/some/file.xlsx', sheet_mapping)

        # Check if load_workbook was called with the correct file
        mock_load_workbook.assert_called_once_with('/some/file.xlsx')

        # Check if the sheet names were correctly renamed
        mock_sheet1.title = 'RenamedSheet1'
        mock_sheet2.title = 'RenamedSheet2'

        # Check if workbook.save was called with the correct file
        mock_workbook.save.assert_called_once_with('/some/file.xlsx')

if __name__ == '__main__':
    unittest.main()
